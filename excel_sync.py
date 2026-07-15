"""投递记录的本地 Excel 镜像。

SQLite 是唯一写入源；此模块只负责把当前投递状态导出为固定工作簿，
避免双向编辑造成数据冲突。
"""

from pathlib import Path
from threading import Lock

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

import paths


MIRROR_FILE = paths.DATA_DIR / "秋招投递追踪.xlsx"
SHEET_NAME = "岗位投递"
HEADERS = [
    ("记录ID", 10), ("公司", 20), ("岗位", 25), ("城市", 14),
    ("当前状态", 16), ("优先级", 10), ("状态更新时间", 20),
    ("下一步行动", 36), ("面试反馈摘要", 44), ("简历路径", 36),
    ("投递来源", 18), ("岗位原始链接", 42), ("JD 原文快照", 70),
]
_LOCK = Lock()


def sync_application_workbook(applications: list[dict] | None = None) -> Path:
    """原子地重建投递镜像，返回工作簿路径。

    Excel 正在打开时 Windows 可能拒绝写入；此时抛出 OSError，由调用方
    记录提示并保留上一次成功写入的工作簿。
    """
    if applications is None:
        import db_manager  # 延迟导入，避免 db_manager 循环依赖
        applications = db_manager.get_applications_with_resume()

    paths.DATA_DIR.mkdir(parents=True, exist_ok=True)
    temp_file = MIRROR_FILE.with_suffix(".tmp.xlsx")
    with _LOCK:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.freeze_panes = "A2"
        last_column = openpyxl.utils.get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{last_column}{max(2, len(applications) + 1)}"

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for column, (title, width) in enumerate(HEADERS, 1):
            cell = ws.cell(1, column, title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width

        for row, app in enumerate(applications, 2):
            values = [
                app.get("id"), app.get("company_name", ""), app.get("position_name", ""),
                app.get("city", ""), app.get("current_status", ""), app.get("priority", 0),
                app.get("status_update_time", ""), app.get("next_action", ""),
                app.get("interview_feedback", ""), app.get("file_path", ""),
                app.get("application_source", ""), app.get("job_link", ""),
                app.get("jd_text", ""),
            ]
            for column, value in enumerate(values, 1):
                cell = ws.cell(row, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=column in (3, 8, 9, 10, 11, 12, 13))
                if column == 12 and value:
                    cell.hyperlink = str(value)
                    cell.style = "Hyperlink"
            ws.row_dimensions[row].height = 48

        if applications:
            table = Table(displayName="Applications", ref=f"A1:{last_column}{len(applications) + 1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(table)

        # 红色标出已终止，绿色标出 Offer；保留普通 Excel 可编辑、可筛选体验。
        ws.conditional_formatting.add(
            f"E2:E{max(2, len(applications) + 1)}",
            FormulaRule(formula=['E2="Offer"'], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            f"E2:E{max(2, len(applications) + 1)}",
            FormulaRule(formula=['E2="终止"'], fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        ws.sheet_view.showGridLines = False
        wb.save(temp_file)
        temp_file.replace(MIRROR_FILE)
    return MIRROR_FILE
