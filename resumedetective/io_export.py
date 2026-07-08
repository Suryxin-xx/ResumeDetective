"""
xlsx 批量导入导出模块
导入格式：公司名 | 岗位名 | 简历文件路径 | JD原文 | 当前状态 | 面试反馈 | 下一步计划 | 备注
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from . import db_manager
from . import paths
from .dialogs import AddResumeDialog

# 导入模板列定义
EXCEL_HEADERS = [
    ("公司名", 20),
    ("岗位名", 20),
    ("简历文件路径", 35),
    ("JD原文", 40),
    ("当前状态", 16),
    ("面试反馈", 40),
    ("下一步计划", 30),
    ("备注", 20),
]

STATUS_VALUES = ["已投递", "简历初筛", "笔试/无笔试", "业务面试", "HR面", "Offer", "终止"]


def _sanitize_sheet_name(name):
    """Excel sheet 名不能超过 31 字符，去掉非法字符"""
    clean = "".join(c for c in name if c not in r"\/*?:[]")
    return clean[:31]


def export_xlsx(output_path):
    """
    导出所有投递数据到 xlsx
    返回导出的记录数
    """
    apps = db_manager.get_applications_with_resume()
    if not apps:
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投递记录"

    # 写表头
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, (name, width) in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 写数据
    for row_idx, app in enumerate(apps, 2):
        ws.cell(row=row_idx, column=1, value=app["company_name"])
        ws.cell(row=row_idx, column=2, value=app["position_name"])
        ws.cell(row=row_idx, column=3, value=app.get("file_path", ""))
        ws.cell(row=row_idx, column=4, value=app.get("jd_text", ""))
        ws.cell(row=row_idx, column=5, value=app["current_status"])
        ws.cell(row=row_idx, column=6, value=app.get("interview_feedback", ""))
        ws.cell(row=row_idx, column=7, value=app.get("next_action", ""))
        ws.cell(row=row_idx, column=8, value="")

    wb.save(output_path)
    return len(apps)


def import_xlsx(input_path):
    """
    从 xlsx 批量导入投递记录
    支持格式：第一行表头，后续每行一条记录

    如果指定了简历文件路径且文件存在，自动拷贝到 data/Resumes/
    返回导入的记录数
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    if ws is None:
        return 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0

    for row in rows:
        if not row or not row[0] or not row[1]:
            continue  # 跳过空行或缺少公司名/岗位名的行

        company = str(row[0]).strip()
        position = str(row[1]).strip()
        if not company or not position:
            continue

        source_file = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        jd_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        status = str(row[4]).strip() if len(row) > 4 and row[4] else "已投递"
        feedback = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        next_action = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        version_note = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        # 校验状态
        if status not in STATUS_VALUES:
            status = "已投递"

        # 拷贝简历文件
        rel_path = ""
        if source_file and Path(source_file).exists():
            rel_path = AddResumeDialog.copy_file_to_resumes(
                source_file, company, position
            )

        # 写入数据库
        rid = db_manager.add_resume(
            company_name=company,
            position_name=position,
            file_path=rel_path,
            jd_text=jd_text,
            version_note=version_note,
        )
        app_id = db_manager.add_application(rid, status)

        # 写入反馈和下一步
        if feedback or next_action:
            conn = db_manager.get_connection()
            try:
                conn.execute(
                    "UPDATE applications SET interview_feedback=?, next_action=? WHERE id=?",
                    (feedback, next_action, app_id),
                )
                conn.commit()
            finally:
                conn.close()

        count += 1

    wb.close()
    return count


def generate_template(output_path):
    """生成空白导入模板 xlsx（带表头说明 + 示例数据 + 状态说明）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投递记录"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_desc = [
        "必填", "必填", "可选，本地简历文件路径", "可选",
        f"可选，默认'已投递'，可选值见「状态说明」sheet", "可选", "可选", "可选",
    ]
    for col_idx, ((name, width), desc) in enumerate(zip(EXCEL_HEADERS, header_desc), 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[chr(64 + col_idx)].width = width
        ws.cell(row=2, column=col_idx, value=desc).font = Font(color="888888", italic=True, size=9)

    example = ["字节跳动", "后端开发", r"C:\Users\xxx\简历.pdf",
               "负责后端服务开发...", "业务面试", "一面通过", "准备二面", "v2.0"]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=3, column=col_idx, value=val)

    ws2 = wb.create_sheet("状态说明")
    ws2.cell(row=1, column=1, value="可选状态值").font = Font(bold=True, size=12)
    for i, s in enumerate(STATUS_VALUES, 2):
        ws2.cell(row=i, column=1, value=s)
    wb.save(output_path)


# ─── 资料库导入导出 ──────────────────────────

MATERIAL_HEADERS = [
    ("类型", 16), ("标题", 30), ("内容", 60), ("标签", 30),
    ("开始时间", 14), ("结束时间", 14),
]
MATERIAL_TYPES = ["项目经历", "实习经历", "获奖经历", "科研项目",
                  "竞赛经历", "学生工作", "志愿活动", "技能证书"]


def export_materials_xlsx(output_path):
    """导出所有经历碎片到 xlsx"""
    materials = db_manager.get_all_materials()
    if not materials:
        return 0
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "经历碎片"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, (name, w) in enumerate(MATERIAL_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font, c.fill = hf, hfill
        ws.column_dimensions[chr(64 + ci)].width = w
    for ri, m in enumerate(materials, 2):
        ws.cell(row=ri, column=1, value=m.get("material_type", ""))
        ws.cell(row=ri, column=2, value=m.get("title", ""))
        ws.cell(row=ri, column=3, value=m.get("content", ""))
        ws.cell(row=ri, column=4, value=m.get("tags", ""))
        ws.cell(row=ri, column=5, value=m.get("start_time", ""))
        ws.cell(row=ri, column=6, value=m.get("end_time", ""))
    wb.save(output_path)
    return len(materials)


def import_materials_xlsx(input_path):
    """从 xlsx 批量导入经历碎片"""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    if ws is None:
        return 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    for row in rows:
        if not row or not row[1]:
            continue
        mtype = str(row[0]).strip() if row[0] else "项目经历"
        title = str(row[1]).strip()
        content = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        tags = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        st = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        et = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        db_manager.add_material(mtype, title, content, tags, st, et)
        count += 1
    wb.close()
    return count


def generate_materials_template(output_path):
    """生成空白经历碎片导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "经历碎片"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    descs = ["必填，可选值见「类型说明」", "必填", "必填", "可选，逗号分隔", "可选 yyyy-MM-dd", "可选 yyyy-MM-dd"]
    for ci, ((name, w), desc) in enumerate(zip(MATERIAL_HEADERS, descs), 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font, c.fill = hf, hfill
        ws.column_dimensions[chr(64 + ci)].width = w
        ws.cell(row=2, column=ci, value=desc).font = Font(color="888888", italic=True, size=9)

    example = ["项目经历", "智能推荐系统", "使用协同过滤算法实现个性化推荐...",
               "推荐系统,Python,机器学习", "2025-01", "2025-06"]
    for ci, v in enumerate(example, 1):
        ws.cell(row=3, column=ci, value=v)

    ws2 = wb.create_sheet("类型说明")
    ws2.cell(row=1, column=1, value="可选类型值").font = Font(bold=True, size=12)
    for i, t in enumerate(MATERIAL_TYPES, 2):
        ws2.cell(row=i, column=1, value=t)
    wb.save(output_path)
